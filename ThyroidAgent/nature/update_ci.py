#!/usr/bin/env python3
"""Update Fig2b-e with CI_low, CI_high from Table S3/S4 ± values."""

import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

OUT = "ThyroidAgent/nature/source_data/Fig2_SegCls_Source_Data.xlsx"

# ── Table S3: Segmentation (± values are CI half-width) ──
seg_datasets = ["TN3K", "ThyroidXL", "PKTN", "TN5K", "DDTI", "ZJH-8K", "RJH-7K"]
seg_models = ["TransUNet", "MedSegX", "MedSAM2", "UltraFedFM", "ThyroidXAgent"]

# (mean, half_width)
dice = {
    "TransUNet":  [(81.84,1.62),(85.75,0.57),(76.89,3.56),(78.54,1.51),(76.58,1.62),(80.72,0.97),(84.83,0.37)],
    "MedSegX":    [(83.93,0.79),(79.98,0.36),(80.63,0.42),(83.10,0.48),(75.12,1.68),(84.06,0.39),(85.40,0.18)],
    "MedSAM2":    [(84.47,1.02),(86.94,0.36),(83.46,2.60),(83.03,1.29),(84.72,1.26),(86.29,0.73),(90.72,0.21)],
    "UltraFedFM": [(81.18,1.46),(84.70,0.53),(75.31,1.12),(77.13,1.38),(75.57,1.67),(80.64,0.84),(83.10,0.33)],
    "ThyroidXAgent":[(85.28,1.28),(87.58,0.44),(82.99,2.10),(83.26,1.34),(85.62,1.07),(94.30,0.38),(91.46,0.14)],
}
hd95 = {
    "TransUNet":  [(27.27,5.52),(22.42,1.34),(26.88,9.66),(22.32,3.43),(17.12,1.55),(18.37,0.75),(18.81,0.74)],
    "MedSegX":    [(10.95,0.64),(11.07,0.32),(10.83,0.70),(11.76,0.76),(18.39,1.65),(10.96,0.35),(9.37,0.18)],
    "MedSAM2":    [(11.51,1.53),(5.46,0.44),(10.56,3.64),(10.94,1.12),(10.06,1.21),(6.79,0.57),(2.92,0.17)],
    "UltraFedFM": [(14.98,2.10),(8.10,0.58),(16.08,1.67),(14.96,1.65),(18.12,1.47),(8.69,0.80),(9.06,0.38)],
    "ThyroidXAgent":[(10.31,1.70),(5.43,0.53),(9.01,3.58),(10.12,1.23),(9.24,1.07),(2.25,0.39),(1.92,0.08)],
}

# ── Table S4: Classification (± values are CI half-width) ──
cls_datasets = ["TN3K", "ThyroidXL", "TN5K", "DDTI", "ZJH-8K"]
cls_models = [
    "ResNet-50","RepViT","LSNet","UltraFedFM","MedGemma",
    "Qwen3-VL-8B","GPT-5","Gemini-2.5-Pro","ThyroidXAgent",
]
auroc = {
    "ResNet-50":      [(0.7674,0.0394),(0.9044,0.0118),(0.9322,0.0168),(0.6704,0.0842),(0.6704,0.0842)],
    "RepViT":         [(0.5556,0.0463),(0.7774,0.0188),(0.6603,0.0375),(0.6162,0.0804),(0.8538,0.0185)],
    "LSNet":          [(0.8095,0.0333),(0.9178,0.0114),(0.9091,0.0201),(0.7581,0.0658),(0.8631,0.0201)],
    "UltraFedFM":     [(0.8461,0.0697),(0.9239,0.0104),(0.9298,0.0175),(0.7518,0.1712),(0.9115,0.0140)],
    "MedGemma":       [(0.8492,0.0305),(0.9371,0.0095),(0.9442,0.0156),(0.8255,0.0650),(0.8976,0.0166)],
    "Qwen3-VL-8B":    [(0.8237,0.0328),(0.9050,0.0115),(0.9214,0.0187),(0.7361,0.0692),(0.8659,0.0189)],
    "GPT-5":          [(0.6924,0.0421),(0.7059,0.0469),(0.7737,0.0996),(0.6346,0.0914),(0.6109,0.0515)],
    "Gemini-2.5-Pro": [(0.6587,0.0455),(0.6246,0.0640),(0.6873,0.0691),(0.6156,0.1308),(0.6493,0.0516)],
    "ThyroidXAgent":  [(0.8692,0.0349),(0.9676,0.0066),(0.9472,0.0152),(0.7991,0.0741),(0.9175,0.0167)],
}
auprc = {
    "ResNet-50":      [(0.6882,0.0632),(0.8882,0.0174),(0.9674,0.0268),(0.3755,0.1176),(0.2755,0.1167)],
    "RepViT":         [(0.4275,0.0528),(0.7161,0.0276),(0.8403,0.0216),(0.3924,0.0933),(0.9486,0.0078)],
    "LSNet":          [(0.7581,0.0452),(0.9040,0.0142),(0.9551,0.0134),(0.4180,0.1410),(0.9449,0.0113)],
    "UltraFedFM":     [(0.8531,0.0284),(0.9354,0.0114),(0.8422,0.0421),(0.4487,0.1452),(0.9669,0.0084)],
    "MedGemma":       [(0.8047,0.0430),(0.9201,0.0139),(0.9747,0.0084),(0.5537,0.1663),(0.9589,0.0096)],
    "Qwen3-VL-8B":    [(0.7617,0.0511),(0.8787,0.0379),(0.9636,0.0106),(0.4112,0.1415),(0.9498,0.0096)],
    "GPT-5":          [(0.6627,0.0633),(0.6237,0.0666),(0.8920,0.0316),(0.3578,0.1089),(0.8311,0.0377)],
    "Gemini-2.5-Pro": [(0.6205,0.0587),(0.4914,0.0841),(0.8462,0.0446),(0.3924,0.1527),(0.8403,0.0362)],
    "ThyroidXAgent":  [(0.8545,0.0600),(0.9653,0.0078),(0.9752,0.0089),(0.5863,0.1380),(0.9711,0.0006)],
}

# ── Helpers ──
header_font = Font(bold=True, size=11)
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

def write_triple(ws, datasets, models, data_dict, metric_name, fig_label):
    """Write a sheet with Mean, CI_low, CI_high per model per dataset."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row+1):
        for cell in row:
            cell.value = None
    ws.delete_rows(1, ws.max_row)

    ws.cell(row=1, column=1).value = f"{fig_label}. {metric_name} across datasets with 95% CI."
    ws.cell(row=2, column=1).value = "CI_low = Mean - half_width; CI_high = Mean + half_width (from bootstrap)."
    ws.cell(row=3, column=1).value = ""

    # Header row 4: Dataset, then Model1_Mean, Model1_CI_low, Model1_CI_high, Model2_Mean, ...
    ws.cell(row=4, column=1).value = "Dataset"
    col = 2
    for m in models:
        ws.cell(row=4, column=col).value = f"{m}_Mean"
        ws.cell(row=4, column=col+1).value = f"{m}_CI_low"
        ws.cell(row=4, column=col+2).value = f"{m}_CI_high"
        col += 3

    for ri, ds in enumerate(datasets):
        r = 5 + ri
        ws.cell(row=r, column=1).value = ds
        col = 2
        for m in models:
            mean_val, hw = data_dict[m][ri]
            ws.cell(row=r, column=col).value = round(mean_val, 6)
            ws.cell(row=r, column=col+1).value = round(mean_val - hw, 6)
            ws.cell(row=r, column=col+2).value = round(mean_val + hw, 6)
            col += 3

    # Style header row
    max_col = 1 + len(models) * 3
    for c in range(1, max_col + 1):
        ws.cell(row=4, column=c).font = header_font
        ws.cell(row=4, column=c).border = thin
    # Auto width
    for col_cells in ws.columns:
        mx = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(mx + 3, 22)

wb = openpyxl.load_workbook(OUT)

write_triple(wb["Fig2b"], seg_datasets, seg_models, dice,   "Nodule Dice (%)",      "Fig. 2b")
write_triple(wb["Fig2c"], seg_datasets, seg_models, hd95,   "Nodule HD95 (mm)",      "Fig. 2c")
write_triple(wb["Fig2d"], cls_datasets, cls_models, auroc,  "Classification AUROC",  "Fig. 2d")
write_triple(wb["Fig2e"], cls_datasets, cls_models, auprc,  "Classification AUPRC",  "Fig. 2e")

wb.save(OUT)
print("Updated Fig2b-e with Mean + CI_low + CI_high from Table S3/S4.")
