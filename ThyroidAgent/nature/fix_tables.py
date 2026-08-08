import openpyxl

wb = openpyxl.load_workbook('ThyroidAgent/nature/source_data/Supplementary_SegCls_Source_Data.xlsx')
ws = wb['TableS3']

ws.cell(row=1).value = 'Table S3. Cross-dataset generalization for thyroid nodule segmentation.'
ws.cell(row=2).value = 'Dice (%) and HD95 (mm) values are point estimates. CI computed via bootstrap.'

ws.cell(row=4, column=1).value = 'Dataset'
ws.cell(row=4, column=2).value = 'Model'
ws.cell(row=4, column=3).value = 'Dice'
ws.cell(row=4, column=4).value = 'HD95'

for r in range(1, ws.max_row + 1):
    for c in [5, 6]:
        ws.cell(row=r, column=c).value = None

print('Fixed:', [ws.cell(row=4, column=c).value for c in range(1, 7)])
print('Sample:', [ws.cell(row=5, column=c).value for c in range(1, 7)])

wb.save('ThyroidAgent/nature/source_data/Supplementary_SegCls_Source_Data.xlsx')
print('Done.')
