import openpyxl
from openpyxl.styles import Font, Border, Side

header_font = Font(bold=True, size=11)
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

wb = openpyxl.load_workbook("ThyroidAgent/nature/source_data/Fig3_Malignant_Source_Data.xlsx")

# Read current Fig3b
ws_old = wb["Fig3b"]
models = []
for r in range(5, ws_old.max_row + 1):
    m = ws_old.cell(row=r, column=1).value
    if not m:
        continue
    vals = [ws_old.cell(row=r, column=c).value for c in range(2, 14)]
    models.append((m, vals))

# Remove old
del wb["Fig3b"]

# --- Fig3b-LNM: only models with LNM_AUROC not N/A ---
lnm = [(m, v) for m, v in models if isinstance(v[0], (int, float))]
ws = wb.create_sheet("Fig3b-LNM")
ws.cell(row=1, column=1).value = "Fig. 3b (LNM). Lateral lymph-node metastasis prediction performance with 95% CI."
ws.cell(row=2, column=1).value = "AUROC and AUPRC reported on the 158-image LymphUs Center 2 test set."
ws.cell(row=3, column=1).value = ""
headers = ["Method", "AUROC", "AUROC_CI_low", "AUROC_CI_high", "AUPRC", "AUPRC_CI_low", "AUPRC_CI_high"]
for ci, h in enumerate(headers, 1):
    ws.cell(row=4, column=ci).value = h
    ws.cell(row=4, column=ci).font = header_font
    ws.cell(row=4, column=ci).border = thin
for i, (m, vals) in enumerate(lnm):
    r = 5 + i
    ws.cell(row=r, column=1).value = m
    ws.cell(row=r, column=2).value = vals[0]   # LNM_AUROC
    ws.cell(row=r, column=3).value = vals[1]   # low
    ws.cell(row=r, column=4).value = vals[2]   # high
    ws.cell(row=r, column=5).value = vals[3]   # LNM_AUPRC
    ws.cell(row=r, column=6).value = vals[4]   # low
    ws.cell(row=r, column=7).value = vals[5]   # high
print("Fig3b-LNM: {} models".format(len(lnm)))

# --- Fig3b-FTCPTC: only models with FTC/PTC_AUROC not N/A ---
ftc = [(m, v) for m, v in models if isinstance(v[6], (int, float))]
ws = wb.create_sheet("Fig3b-FTCPTC")
ws.cell(row=1, column=1).value = "Fig. 3b (FTC/PTC). Follicular vs papillary thyroid carcinoma classification performance with 95% CI."
ws.cell(row=2, column=1).value = "AUROC and AUPRC reported on the 200-image Dai et al. test set."
ws.cell(row=3, column=1).value = ""
for ci, h in enumerate(headers, 1):
    ws.cell(row=4, column=ci).value = h
    ws.cell(row=4, column=ci).font = header_font
    ws.cell(row=4, column=ci).border = thin
for i, (m, vals) in enumerate(ftc):
    r = 5 + i
    ws.cell(row=r, column=1).value = m
    ws.cell(row=r, column=2).value = vals[6]   # FTC_AUROC
    ws.cell(row=r, column=3).value = vals[7]
    ws.cell(row=r, column=4).value = vals[8]
    ws.cell(row=r, column=5).value = vals[9]   # FTC_AUPRC
    ws.cell(row=r, column=6).value = vals[10]
    ws.cell(row=r, column=7).value = vals[11]
print("Fig3b-FTCPTC: {} models".format(len(ftc)))

wb.save("ThyroidAgent/nature/source_data/Fig3_Malignant_Source_Data.xlsx")
print("Done. Sheets: {}".format(wb.sheetnames))
