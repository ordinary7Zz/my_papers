import openpyxl
from openpyxl.styles import Font, Border, Side
from collections import Counter

header_font = Font(bold=True, size=11)
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

wb = openpyxl.load_workbook("ThyroidAgent/nature/source_data/Supplementary_SegCls_Source_Data.xlsx")

# --- FigS2_centroid_kde ---
ws = wb["FigS2_centroid_kde"]
rows_data = []
for r in range(5, ws.max_row + 1):
    ds = ws.cell(row=r, column=1).value
    cx = ws.cell(row=r, column=2).value
    cy = ws.cell(row=r, column=3).value
    if ds and cx is not None:
        rows_data.append((ds, cx, cy))

counters = Counter()
for row in ws.iter_rows(min_row=1, max_row=ws.max_row + 1):
    for cell in row:
        cell.value = None
ws.delete_rows(1, ws.max_row)

ws.cell(row=1, column=1).value = "Fig. S2 (bottom-left). Normalized lesion-mask centroid positions across datasets."
ws.cell(row=2, column=1).value = "x and y are normalized to [0,1] within the ultrasound frame."
ws.cell(row=4, column=1).value = "Dataset"
ws.cell(row=4, column=2).value = "Case_ID"
ws.cell(row=4, column=3).value = "Centroid_X"
ws.cell(row=4, column=4).value = "Centroid_Y"
for c in range(1, 5):
    ws.cell(row=4, column=c).font = header_font
    ws.cell(row=4, column=c).border = thin

for i, (ds, cx, cy) in enumerate(rows_data):
    counters[ds] += 1
    cid = "Case_{:04d}".format(counters[ds])
    r = 5 + i
    ws.cell(row=r, column=1).value = ds
    ws.cell(row=r, column=2).value = cid
    ws.cell(row=r, column=3).value = cx
    ws.cell(row=r, column=4).value = cy
print("centroid_kde: {} rows with Case_ID".format(ws.max_row))

# --- FigS2_size_kde ---
ws = wb["FigS2_size_kde"]
rows_data2 = []
for r in range(5, ws.max_row + 1):
    ds = ws.cell(row=r, column=1).value
    rs = ws.cell(row=r, column=2).value
    if ds and rs is not None:
        rows_data2.append((ds, rs))

counters2 = Counter()
for row in ws.iter_rows(min_row=1, max_row=ws.max_row + 1):
    for cell in row:
        cell.value = None
ws.delete_rows(1, ws.max_row)

ws.cell(row=1, column=1).value = "Fig. S2 (bottom-right). Relative lesion size distributions across datasets."
ws.cell(row=2, column=1).value = "relative_size = mask_area / image_area."
ws.cell(row=4, column=1).value = "Dataset"
ws.cell(row=4, column=2).value = "Case_ID"
ws.cell(row=4, column=3).value = "Relative_Size"
for c in range(1, 4):
    ws.cell(row=4, column=c).font = header_font
    ws.cell(row=4, column=c).border = thin

for i, (ds, rs) in enumerate(rows_data2):
    counters2[ds] += 1
    cid = "Case_{:04d}".format(counters2[ds])
    r = 5 + i
    ws.cell(row=r, column=1).value = ds
    ws.cell(row=r, column=2).value = cid
    ws.cell(row=r, column=3).value = rs
print("size_kde: {} rows with Case_ID".format(ws.max_row))

wb.save("ThyroidAgent/nature/source_data/Supplementary_SegCls_Source_Data.xlsx")

# Verify
wb2 = openpyxl.load_workbook("ThyroidAgent/nature/source_data/Supplementary_SegCls_Source_Data.xlsx")
for name in ["FigS2_centroid_kde", "FigS2_size_kde"]:
    ws = wb2[name]
    hdr = [ws.cell(row=4, column=c).value for c in range(1, 5)]
    r5 = [ws.cell(row=5, column=c).value for c in range(1, 5)]
    r_end = [ws.cell(row=ws.max_row, column=c).value for c in range(1, 5)]
    print("{}: hdr={}, R5={}, R{}={}".format(name, hdr, r5, ws.max_row, r_end))
