#!/usr/bin/env python3
"""Update Source Data xlsx with real numbers from Table S3 (seg) and Table S4 (cls)."""

import os
import re
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(__file__), "source_data")

# ── Style ──
header_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.border = thin_border


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ============================================================
# Real data from manuscript Table S3 (seg_performance)
# ============================================================
seg_datasets = ["TN3K", "ThyroidXL", "PKTN", "TN5K", "DDTI", "ZJH-8K", "RJH-7K"]
seg_models = ["TransUNet", "MedSegX", "MedSAM2", "UltraFedFM", "ThyroidXAgent"]

dice_data = {
    "TransUNet":  [81.84, 85.75, 76.89, 78.54, 76.58, 80.72, 84.83],
    "MedSegX":    [83.93, 79.98, 80.63, 83.10, 75.12, 84.06, 85.40],
    "MedSAM2":    [84.47, 86.94, 83.46, 83.03, 84.72, 86.29, 90.72],
    "UltraFedFM": [81.18, 84.70, 75.31, 77.13, 75.57, 80.64, 83.10],
    "ThyroidXAgent": [85.28, 87.58, 82.99, 83.26, 85.62, 94.30, 91.46],
}

hd95_data = {
    "TransUNet":  [27.27, 22.42, 26.88, 22.32, 17.12, 18.37, 18.81],
    "MedSegX":    [10.95, 11.07, 10.83, 11.76, 18.39, 10.96,  9.37],
    "MedSAM2":    [11.51,  5.46, 10.56, 10.94, 10.06,  6.79,  2.92],
    "UltraFedFM": [14.98,  8.10, 16.08, 14.96, 18.12,  8.69,  9.06],
    "ThyroidXAgent": [10.31,  5.43,  9.01, 10.12,  9.24,  2.25,  1.92],
}

# ============================================================
# Real data from manuscript Table S4 (cls_performance)
# ============================================================
cls_datasets = ["TN3K", "ThyroidXL", "TN5K", "DDTI", "ZJH-8K"]
cls_models = [
    "ResNet-50", "RepViT", "LSNet", "UltraFedFM", "MedGemma",
    "Qwen3-VL-8B", "GPT-5", "Gemini-2.5-Pro", "ThyroidXAgent",
]

auroc_data = {
    "ResNet-50":       [0.7674, 0.9044, 0.9322, 0.6704, 0.6704],
    "RepViT":          [0.5556, 0.7774, 0.6603, 0.6162, 0.8538],
    "LSNet":           [0.8095, 0.9178, 0.9091, 0.7581, 0.8631],
    "UltraFedFM":      [0.8461, 0.9239, 0.9298, 0.7518, 0.9115],
    "MedGemma":        [0.8492, 0.9371, 0.9442, 0.8255, 0.8976],
    "Qwen3-VL-8B":     [0.8237, 0.9050, 0.9214, 0.7361, 0.8659],
    "GPT-5":           [0.6924, 0.7059, 0.7737, 0.6346, 0.6109],
    "Gemini-2.5-Pro":  [0.6587, 0.6246, 0.6873, 0.6156, 0.6493],
    "ThyroidXAgent":   [0.8692, 0.9676, 0.9472, 0.7991, 0.9175],
}

auprc_data = {
    "ResNet-50":       [0.6882, 0.8882, 0.9674, 0.3755, 0.2755],
    "RepViT":          [0.4275, 0.7161, 0.8403, 0.3924, 0.9486],
    "LSNet":           [0.7581, 0.9040, 0.9551, 0.4180, 0.9449],
    "UltraFedFM":      [0.8531, 0.9354, 0.8422, 0.4487, 0.9669],
    "MedGemma":        [0.8047, 0.9201, 0.9747, 0.5537, 0.9589],
    "Qwen3-VL-8B":     [0.7617, 0.8787, 0.9636, 0.4112, 0.9498],
    "GPT-5":           [0.6627, 0.6237, 0.8920, 0.3578, 0.8311],
    "Gemini-2.5-Pro":  [0.6205, 0.4914, 0.8462, 0.3924, 0.8403],
    "ThyroidXAgent":   [0.8545, 0.9653, 0.9752, 0.5863, 0.9711],
}


# ============================================================
# Update File 1: Fig2_SegCls_Source_Data.xlsx
# ============================================================
wb = openpyxl.load_workbook(f"{OUT_DIR}/Fig2_SegCls_Source_Data.xlsx")

# --- Fig2b: Dice (from Table S3) ---
ws = wb["Fig2b"]
# Clear existing data rows (keep Row 1-3 description, replace from Row 4)
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None
ws.delete_rows(4, ws.max_row - 3)

# Rebuild with real data
# Update models in header
for ci, m in enumerate(seg_models, 2):
    ws.cell(row=3, column=ci).value = m
for ri, ds in enumerate(seg_datasets, 4):
    ws.cell(row=ri, column=1).value = ds
    for ci, m in enumerate(seg_models, 2):
        ws.cell(row=ri, column=ci).value = dice_data[m][ri - 4]
style_header(ws, 3, len(seg_models) + 1)
auto_width(ws)

# --- Fig2c: HD95 (from Table S3) ---
ws = wb["Fig2c"]
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None
ws.delete_rows(4, ws.max_row - 3)
for ci, m in enumerate(seg_models, 2):
    ws.cell(row=3, column=ci).value = m
for ri, ds in enumerate(seg_datasets, 4):
    ws.cell(row=ri, column=1).value = ds
    for ci, m in enumerate(seg_models, 2):
        ws.cell(row=ri, column=ci).value = hd95_data[m][ri - 4]
style_header(ws, 3, len(seg_models) + 1)
auto_width(ws)

# --- Fig2d: AUROC (from Table S4) ---
ws = wb["Fig2d"]
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None
ws.delete_rows(4, ws.max_row - 3)
for ci, m in enumerate(cls_models, 2):
    ws.cell(row=3, column=ci).value = m
for ri, ds in enumerate(cls_datasets, 4):
    ws.cell(row=ri, column=1).value = ds
    for ci, m in enumerate(cls_models, 2):
        ws.cell(row=ri, column=ci).value = auroc_data[m][ri - 4]
style_header(ws, 3, len(cls_models) + 1)
auto_width(ws)

# --- Fig2e: AUPRC (from Table S4) ---
ws = wb["Fig2e"]
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None
ws.delete_rows(4, ws.max_row - 3)
for ci, m in enumerate(cls_models, 2):
    ws.cell(row=3, column=ci).value = m
for ri, ds in enumerate(cls_datasets, 4):
    ws.cell(row=ri, column=1).value = ds
    for ci, m in enumerate(cls_models, 2):
        ws.cell(row=ri, column=ci).value = auprc_data[m][ri - 4]
style_header(ws, 3, len(cls_models) + 1)
auto_width(ws)

wb.save(f"{OUT_DIR}/Fig2_SegCls_Source_Data.xlsx")
print(f"Updated Fig2_SegCls_Source_Data.xlsx with real data from Table S3 & S4")


# ============================================================
# Update File 2: Supplementary_SegCls_Source_Data.xlsx
#   – TableS3 and TableS4 sheets
# ============================================================
wb2 = openpyxl.load_workbook(f"{OUT_DIR}/Supplementary_SegCls_Source_Data.xlsx")

# --- TableS3: Segmentation performance (Dice + HD95) ---
ws = wb2["TableS3"]
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None
ws.delete_rows(4, ws.max_row - 3)

# Update header model names
seg_models_renamed = {
    "TransUNet": "TransUNet",
    "MedSegX": "MedSegX",
    "MedSAM2": "MedSAM2",
    "UltraFedFM": "UltraFedFM",
    "ThyroidXAgent": "ThyroidXAgent",
}
ri = 4
for ds in seg_datasets:
    for m in seg_models:
        ws.cell(row=ri, column=1).value = ds
        ws.cell(row=ri, column=2).value = m
        ws.cell(row=ri, column=3).value = dice_data[m][seg_datasets.index(ds)]
        ws.cell(row=ri, column=4).value = hd95_data[m][seg_datasets.index(ds)]
        ws.cell(row=ri, column=5).value = ""   # Gland_Dice placeholder
        ws.cell(row=ri, column=6).value = ""   # Gland_HD95 placeholder
        ri += 1
style_header(ws, 4, 6)
auto_width(ws)

# --- TableS4: Classification performance (AUROC + AUPRC) ---
ws = wb2["TableS4"]
for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
    for cell in row:
        cell.value = None
ws.delete_rows(4, ws.max_row - 3)

ri = 4
for ds in cls_datasets:
    for m in cls_models:
        ws.cell(row=ri, column=1).value = ds
        ws.cell(row=ri, column=2).value = m
        ws.cell(row=ri, column=3).value = auroc_data[m][cls_datasets.index(ds)]
        ws.cell(row=ri, column=4).value = auprc_data[m][cls_datasets.index(ds)]
        ri += 1
style_header(ws, 4, 4)
auto_width(ws)

wb2.save(f"{OUT_DIR}/Supplementary_SegCls_Source_Data.xlsx")
print(f"Updated Supplementary_SegCls_Source_Data.xlsx with real data from Table S3 & S4")

print(f"\nDone. Files in: {OUT_DIR}/")
