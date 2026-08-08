#!/usr/bin/env python3
"""Generate Fig3_Malignant_Source_Data.xlsx from Table S5 and SHAP CSVs."""

import csv, openpyxl, os
from openpyxl.styles import Font, Border, Side

OUT_DIR = os.path.join(os.path.dirname(__file__), "source_data")
os.makedirs(OUT_DIR, exist_ok=True)

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, size=11)
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))

# ────────────────────────────────────────
# Fig3b: Performance from Table S5
# ────────────────────────────────────────
ws = wb.create_sheet("Fig3b")
ws.cell(row=1, column=1).value = "Fig. 3b. Performance on malignant thyroid lesion stratification tasks."
ws.cell(row=2, column=1).value = "Lateral lymph-node metastasis (LNM) and FTC/PTC subtype classification. AUROC and AUPRC with 95% CI."
ws.cell(row=3, column=1).value = ""

table_s5_data = [
    # (Method, LNM_AUROC, LNM_AUROC_low, LNM_AUROC_high, LNM_AUPRC, LNM_AUPRC_low, LNM_AUPRC_high,
    #  FTC_AUROC, FTC_AUROC_low, FTC_AUROC_high, FTC_AUPRC, FTC_AUPRC_low, FTC_AUPRC_high)
    ("RepViT",         0.7905, 0.7229, 0.8581, 0.8152, 0.7514, 0.8790, 0.6419, 0.5580, 0.7258, 0.6297, 0.5355, 0.7239),
    ("LSNet",          0.5878, 0.5013, 0.6743, 0.6301, 0.5426, 0.7176, 0.4858, 0.3933, 0.5783, 0.4845, 0.3937, 0.5753),
    ("UltraFedFM",     0.7757, 0.7026, 0.8488, 0.7902, 0.7057, 0.8747, 0.7365, 0.6621, 0.8109, 0.7582, 0.6758, 0.8406),
    ("MedGemma",       0.8403, 0.7942, 0.8864, 0.8585, 0.8019, 0.9151, 0.6598, 0.5774, 0.7422, 0.6142, 0.5086, 0.7198),
    ("Qwen3-VL-8B",    0.8070, 0.7438, 0.8702, 0.8055, 0.7255, 0.8855, 0.6056, 0.5190, 0.6922, 0.5539, 0.4421, 0.6657),
    ("GPT-5",          0.8410, 0.7835, 0.8985, 0.8629, 0.8096, 0.9162, 0.1604, 0.0898, 0.2310, 0.3638, 0.2791, 0.4485),
    ("Gemini-2.5-Pro", 0.5414, 0.4678, 0.6150, 0.5492, 0.4577, 0.6407, 0.3324, 0.2452, 0.4196, 0.4187, 0.3350, 0.5024),
    ("LLNM-Net",       0.7665, 0.6973, 0.8357, 0.7363, 0.6514, 0.8212, None,  None,  None,  None,  None,  None),
    ("Tiger-Model",    None,  None,  None,  None,  None,  None,  0.7136, 0.6322, 0.7950, 0.7117, 0.6016, 0.8218),
    ("ThyroidXAgent",  0.8642, 0.8092, 0.9192, 0.8808, 0.8271, 0.9345, 0.8053, 0.7454, 0.8652, 0.7863, 0.7070, 0.8656),
]

ws.cell(row=4, column=1).value = "Method"
ws.cell(row=4, column=2).value = "LNM_AUROC"
ws.cell(row=4, column=3).value = "LNM_AUROC_CI_low"
ws.cell(row=4, column=4).value = "LNM_AUROC_CI_high"
ws.cell(row=4, column=5).value = "LNM_AUPRC"
ws.cell(row=4, column=6).value = "LNM_AUPRC_CI_low"
ws.cell(row=4, column=7).value = "LNM_AUPRC_CI_high"
ws.cell(row=4, column=8).value = "FTC/PTC_AUROC"
ws.cell(row=4, column=9).value = "FTC/PTC_AUROC_CI_low"
ws.cell(row=4, column=10).value = "FTC/PTC_AUROC_CI_high"
ws.cell(row=4, column=11).value = "FTC/PTC_AUPRC"
ws.cell(row=4, column=12).value = "FTC/PTC_AUPRC_CI_low"
ws.cell(row=4, column=13).value = "FTC/PTC_AUPRC_CI_high"

for ri, row_data in enumerate(table_s5_data):
    r = 5 + ri
    for ci, val in enumerate(row_data):
        if val is not None:
            ws.cell(row=r, column=ci+1).value = val
        else:
            ws.cell(row=r, column=ci+1).value = "N/A"

for c in range(1, 14):
    ws.cell(row=4, column=c).font = header_font
    ws.cell(row=4, column=c).border = thin

# auto width
for col_cells in ws.columns:
    mx = max((len(str(c.value or "")) for c in col_cells), default=0)
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(mx+3, 22)

# ────────────────────────────────────────
# Fig3c: LNM global SHAP feature importance
# ────────────────────────────────────────
ws = wb.create_sheet("Fig3c-LNM-SHAP")
ws.cell(row=1, column=1).value = "Fig. 3c. Global SHAP feature importance for LNM prediction."
ws.cell(row=2, column=1).value = "Features ranked by mean absolute SHAP value."

with open("results/LNM.csv") as f:
    lnm_feats = [(row["feature"], float(row["mean_abs_shap"])) for row in csv.DictReader(f)]
lnm_feats.sort(key=lambda x: x[1], reverse=True)

ws.cell(row=4, column=1).value = "Rank"
ws.cell(row=4, column=2).value = "Feature"
ws.cell(row=4, column=3).value = "Mean_Abs_SHAP"
for c in range(1, 4):
    ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).border = thin
for i, (feat, val) in enumerate(lnm_feats):
    ws.cell(row=5+i, column=1).value = i+1
    ws.cell(row=5+i, column=2).value = feat
    ws.cell(row=5+i, column=3).value = round(val, 6)

# ────────────────────────────────────────
# Fig3e: FTC/PTC global SHAP feature importance
# ────────────────────────────────────────
ws = wb.create_sheet("Fig3e-FTCPTC-SHAP")
ws.cell(row=1, column=1).value = "Fig. 3e. Global SHAP feature importance for FTC/PTC subtype classification."
ws.cell(row=2, column=1).value = "Features ranked by mean absolute SHAP value."

with open("results/FTCPTC.csv") as f:
    ftc_feats = [(row["feature"], float(row["mean_abs_shap"])) for row in csv.DictReader(f)]
ftc_feats.sort(key=lambda x: x[1], reverse=True)

ws.cell(row=4, column=1).value = "Rank"
ws.cell(row=4, column=2).value = "Feature"
ws.cell(row=4, column=3).value = "Mean_Abs_SHAP"
for c in range(1, 4):
    ws.cell(row=4, column=c).font = header_font; ws.cell(row=4, column=c).border = thin
for i, (feat, val) in enumerate(ftc_feats):
    ws.cell(row=5+i, column=1).value = i+1
    ws.cell(row=5+i, column=2).value = feat
    ws.cell(row=5+i, column=3).value = round(val, 6)

OUT = os.path.join(OUT_DIR, "Fig3_Malignant_Source_Data.xlsx")
wb.save(OUT)
print(f"Created: Fig3_Malignant_Source_Data.xlsx ({len(wb.sheetnames)} sheets)")
for n in wb.sheetnames:
    ws = wb[n]
    print(f"  {n}: {ws.max_row} rows x {ws.max_column} cols")
